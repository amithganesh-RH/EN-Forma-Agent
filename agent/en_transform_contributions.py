#!/usr/bin/env python3
"""
Generates the Forma Contributions upload file from the pretax elections CSV.

Output filename: YYYYMMDD_pretax_contributions.csv

Columns:
  Employee ID, Employee Email, Account Type,
  Employee Deposit Amount, Employer Deposit Amount, Funding Date

Funding Date rules
  1st Monday of month (date 1–7):   previous Friday  (last working day)
  3rd Monday of month (date 15–21): previous Friday
  Both runs are on Mondays so funding date = run_date − 3 days.
"""

import csv
import os
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import List

GDRIVE_FOLDER_ID = os.environ.get("GDRIVE_FOLDER_ID", "1L4qIfD4bha6oZpZTe7s5LbNcgRAgqLlL")
GOOGLE_CREDS     = os.environ.get("GOOGLE_CREDS_FILE", str(Path.home() / ".config/en_report/google_creds.json"))
TOKEN_FILE       = str(Path(GOOGLE_CREDS).parent / "google_token.json")
SCOPES           = [
    "https://www.googleapis.com/auth/drive.file",
]

# Employee ID may have leading zeros (e.g. 000148) — format as Text in Sheets
TEXT_COLUMNS = {"Employee ID"}

OUTPUT_HEADERS = [
    "Employee ID",
    "Employee Email",
    "Account Type",
    "Employee Deposit Amount",
    "Employer Deposit Amount",
    "Funding Date",
]


# ── Funding date ──────────────────────────────────────────────────────────────

def calc_funding_date(today: date) -> str:
    """
    Return the funding date as YYYY-MM-DD.

    For both 1st Monday and 3rd Monday runs the last working day before the
    report is the Friday of the prior week (today − 3 days).
    For any other weekday, walk back to the most recent Friday.
    """
    # weekday(): Mon=0 … Fri=4 … Sun=6
    days_back = (today.weekday() - 4) % 7   # 0 when today IS Friday
    if days_back == 0:
        return today.strftime("%Y-%m-%d")
    return (today - timedelta(days=days_back)).strftime("%Y-%m-%d")


# ── Transform ─────────────────────────────────────────────────────────────────

def transform(elections_csv: Path, today: date) -> List[dict]:
    """
    Read the pretax elections CSV and produce contribution rows.

    Rules:
    - Skip rows with Account Status = Terminated (no deposit needed)
    - Employee Deposit Amount  = Employee Pay Period Election
    - Employer Deposit Amount  = Employer Pay Period Election
    - Funding Date             = calc_funding_date(today)
    """
    funding = calc_funding_date(today)
    rows = []

    with open(elections_csv, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get("Account Status", "").strip() == "Terminated":
                continue

            rows.append({
                "Employee ID":             row.get("Employee ID", "").strip(),
                "Employee Email":          row.get("Employee Email", "").strip(),
                "Account Type":            row.get("Account Type", "").strip(),
                "Employee Deposit Amount": row.get("Employee Pay Period Election", "$0.00").strip(),
                "Employer Deposit Amount": row.get("Employer Pay Period Election", "$0.00").strip(),
                "Funding Date":            funding,
            })

    return rows


# ── Google auth & Drive helpers ───────────────────────────────────────────────

def _get_creds():
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request

    creds = None
    if os.path.exists(TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(GOOGLE_CREDS, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, "w") as f:
            f.write(creds.to_json())
    return creds


def get_drive_service():
    from googleapiclient.discovery import build
    return build("drive", "v3", credentials=_get_creds())


def get_or_create_subfolder(service, parent_id: str, folder_name: str) -> str:
    query = (
        f"name='{folder_name}' and "
        f"'{parent_id}' in parents and "
        "mimeType='application/vnd.google-apps.folder' and "
        "trashed=false"
    )
    results = service.files().list(q=query, fields="files(id, name)").execute()
    files = results.get("files", [])
    if files:
        return files[0]["id"]
    metadata = {
        "name": folder_name,
        "mimeType": "application/vnd.google-apps.folder",
        "parents": [parent_id],
    }
    folder = service.files().create(body=metadata, fields="id").execute()
    return folder["id"]


def save_as_xlsx(rows: List[dict], xlsx_path: Path):
    """Save rows as Excel (.xlsx) with ALL columns set to Text format to prevent auto-conversion."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for col_idx, header in enumerate(OUTPUT_HEADERS, 1):
        ws.cell(row=1, column=col_idx, value=header)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, header in enumerate(OUTPUT_HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(row.get(header, "")))
            cell.number_format = "@"  # Text format on ALL columns
    wb.save(str(xlsx_path))


def xlsx_to_csv(xlsx_path: Path, csv_path: Path):
    """
    Read the saved XLSX (all cells stored as text) and export as UTF-8 BOM CSV.
    Values are taken directly from Excel's stored cell values so leading zeros,
    employee IDs, etc. are preserved exactly as they appear in the Excel file.
    """
    from openpyxl import load_workbook
    wb = load_workbook(str(xlsx_path), data_only=True)
    ws = wb.active
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([str(cell) if cell is not None else "" for cell in row])


def upload_file(drive_service, file_path: Path, folder_id: str, mime_type: str) -> str:
    """Upload a file to Drive, replacing any existing file with the same name."""
    from googleapiclient.http import MediaFileUpload
    media = MediaFileUpload(str(file_path), mimetype=mime_type, resumable=True)
    existing = drive_service.files().list(
        q=f"name='{file_path.name}' and '{folder_id}' in parents and trashed=false",
        fields="files(id)"
    ).execute().get("files", [])
    if existing:
        uploaded = drive_service.files().update(
            fileId=existing[0]["id"], media_body=media, fields="id, name, webViewLink"
        ).execute()
    else:
        uploaded = drive_service.files().create(
            body={"name": file_path.name, "parents": [folder_id]},
            media_body=media, fields="id, name, webViewLink"
        ).execute()
    return uploaded.get("webViewLink", "")


# ── Entry point ───────────────────────────────────────────────────────────────

def run(elections_csv: Path, output_dir: Path, date_str: str = None, today: date = None) -> str:
    """
    Build contributions file from elections CSV, save as Excel and CSV,
    upload XLSX to Drive and CSV to a 'CSV' subfolder. Returns the XLSX Drive link.
    """
    if date_str is None:
        date_str = datetime.now().strftime("%Y%m%d")
    if today is None:
        today = datetime.now().date()

    print(f"  [contributions] Processing elections data...")
    rows = transform(elections_csv, today)
    print(f"  [contributions] {len(rows)} contribution rows (active employees)")

    # 1. Save as XLSX — all columns set to Text so nothing is auto-converted
    xlsx_filename = f"{date_str}_pretax_contributions.xlsx"
    xlsx_path     = output_dir / xlsx_filename
    save_as_xlsx(rows, xlsx_path)
    print(f"  [contributions] Saved: {xlsx_filename}  (funding date: {calc_funding_date(today)})")

    # 2. Derive CSV from the Excel file (values read back exactly as stored)
    csv_filename  = f"{date_str}_pretax_contributions.csv"
    csv_path      = output_dir / csv_filename
    xlsx_to_csv(xlsx_path, csv_path)
    print(f"  [contributions] Saved CSV from Excel: {csv_filename}")

    # 3. Upload to Drive
    print(f"  [contributions] Uploading to Google Drive subfolder...")
    drive_svc    = get_drive_service()
    XLSX_MIME    = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    subfolder_id = get_or_create_subfolder(drive_svc, GDRIVE_FOLDER_ID, "Contributions Upload File")
    xlsx_link    = upload_file(drive_svc, xlsx_path, subfolder_id, XLSX_MIME)
    print(f"  [contributions] ✓ Uploaded: {xlsx_filename}  →  {xlsx_link}")

    csv_sub_id   = get_or_create_subfolder(drive_svc, subfolder_id, "CSV")
    csv_link     = upload_file(drive_svc, csv_path, csv_sub_id, "text/csv")
    print(f"  [contributions] ✓ Uploaded CSV: {csv_filename}  →  {csv_link}")

    return xlsx_link


if __name__ == "__main__":
    import sys
    today_date = datetime.now().date()
    date_str   = datetime.now().strftime("%Y%m%d")

    if len(sys.argv) >= 2:
        elections_path = Path(sys.argv[1])
    else:
        elections_path = Path.home() / "employee_navigator_reports" / f"{date_str}_pretax_elections.csv"

    if not elections_path.exists():
        print(f"ERROR: File not found: {elections_path}")
        sys.exit(1)

    output_dir = Path.home() / "employee_navigator_reports"
    run(elections_path, output_dir, date_str, today_date)
