#!/usr/bin/env python3
"""
Transforms the raw Employee Navigator Demographics CSV into the
pretax demographics format and uploads it to a Google Drive subfolder.

Output filename: YYYYMMDD_pretax_demographics.csv

The file is uploaded to Drive as a Google Sheet with Employee ID, ZIP Code,
and SSN columns formatted as Text, so leading zeros are never stripped when
opened in Google Sheets or Excel. Download as CSV from Sheets to upload to Forma.
"""

import csv
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import List

GDRIVE_FOLDER_ID = os.environ.get("GDRIVE_FOLDER_ID", "1L4qIfD4bha6oZpZTe7s5LbNcgRAgqLlL")
GOOGLE_CREDS     = os.environ.get("GOOGLE_CREDS_FILE", str(Path.home() / ".config/en_report/google_creds.json"))
TOKEN_FILE       = str(Path(GOOGLE_CREDS).parent / "google_token.json")
SCOPES           = [
    "https://www.googleapis.com/auth/drive.file",
]

# These columns contain values with leading zeros that must not be auto-formatted as numbers
TEXT_COLUMNS = {"Employee ID", "ZIP Code", "SSN"}

OUTPUT_HEADERS = [
    "Employee ID",
    "Employee Email",
    "Country",
    "Legal Name - First Name",
    "Legal Name - Last Name",
    "Employee Status",
    "Date of Birth",
    "Employee Personal Email",
    "Preferred Name - First Name",
    "Preferred Name - Last Name",
    "Address Line 1",
    "Address Line 2",
    "City",
    "State",
    "ZIP Code",
    "Phone",
    "Hire Date",
    "Termination Date",
    "HDHP",
    "SSN",
]


def fix_zip(zip_code: str) -> str:
    """Pad 4-digit zip base with a leading 0 (e.g. '1234' → '01234', '1234-5678' → '01234-5678')."""
    z = zip_code.strip()
    if "-" in z:
        base, suffix = z.split("-", 1)
        if len(base) == 4 and base.isdigit():
            base = "0" + base
        return f"{base}-{suffix}"
    if len(z) == 4 and z.isdigit():
        z = "0" + z
    return z


def fix_employee_id(emp_id: str, first: str, last: str) -> str:
    """Ensure Brett Kempker's Employee ID always starts with '000'."""
    if first.strip().lower() == "brett" and last.strip().lower() == "kempker":
        if not emp_id.startswith("000"):
            emp_id = "000" + emp_id
    return emp_id


def transform(demographics_csv: Path) -> List[dict]:
    """Read demographics CSV, deduplicate by Employee ID, map to output format."""
    seen = set()
    rows = []

    with open(demographics_csv, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            emp_id = row.get("Employee ID", "").strip()
            if not emp_id or emp_id in seen:
                continue
            seen.add(emp_id)

            first = row.get("First Name", "").strip()
            last  = row.get("Last Name", "").strip()
            termination_date = row.get("Termination Date", "").strip()
            status = "Terminated" if termination_date else "Active"

            rows.append({
                "Employee ID":               fix_employee_id(emp_id, first, last),
                "Employee Email":            row.get("Work Email", "").strip(),
                "Country":                   row.get("Country", "").strip(),
                "Legal Name - First Name":   first,
                "Legal Name - Last Name":    last,
                "Employee Status":           status,
                "Date of Birth":             row.get("DOB", "").strip(),
                "Employee Personal Email":   row.get("Personal Email", "").strip(),
                "Preferred Name - First Name": first,
                "Preferred Name - Last Name":  last,
                "Address Line 1":            row.get("Address 1", "").strip(),
                "Address Line 2":            row.get("Address 2", "").strip(),
                "City":                      row.get("City", "").strip(),
                "State":                     row.get("State", "").strip(),
                "ZIP Code":                  fix_zip(row.get("Zip", "")),
                "Phone":                     row.get("Mobile Phone", "").strip(),
                "Hire Date":                 row.get("Hire Date", "").strip(),
                "Termination Date":          termination_date,
                "HDHP":                      "TRUE",
                "SSN":                       row.get("Social Security Number", "").strip(),
            })

    return rows


def _get_creds():
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request

    creds = None
    if os.path.exists(TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(GOOGLE_CREDS, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, "w") as f:
            f.write(creds.to_json())
    return creds


def get_drive_service():
    from googleapiclient.discovery import build
    return build("drive", "v3", credentials=_get_creds())


def get_or_create_subfolder(service, parent_id: str, folder_name: str) -> str:
    """Return existing subfolder ID or create it."""
    query = (
        f"name='{folder_name}' and "
        f"'{parent_id}' in parents and "
        "mimeType='application/vnd.google-apps.folder' and "
        "trashed=false"
    )
    results = service.files().list(q=query, fields="files(id, name)").execute()
    files = results.get("files", [])
    if files:
        return files[0]["id"]

    metadata = {
        "name": folder_name,
        "mimeType": "application/vnd.google-apps.folder",
        "parents": [parent_id],
    }
    folder = service.files().create(body=metadata, fields="id").execute()
    return folder["id"]


def save_as_xlsx(rows: List[dict], xlsx_path: Path):
    """
    Save rows as Excel (.xlsx) with ALL columns set to Text format so no value
    (Employee ID, ZIP Code, SSN, dollar amounts, etc.) is auto-converted by Excel.
    """
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for col_idx, header in enumerate(OUTPUT_HEADERS, 1):
        ws.cell(row=1, column=col_idx, value=header)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, header in enumerate(OUTPUT_HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(row.get(header, "")))
            cell.number_format = "@"  # Text format on ALL columns
    wb.save(str(xlsx_path))


def xlsx_to_csv(xlsx_path: Path, csv_path: Path):
    """
    Read the saved XLSX (all cells stored as text) and export as UTF-8 BOM CSV.
    Values are taken directly from Excel's stored cell values so leading zeros,
    SSNs, ZIP codes, etc. are preserved exactly as they appear in the Excel file.
    """
    from openpyxl import load_workbook
    wb = load_workbook(str(xlsx_path), data_only=True)
    ws = wb.active
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([str(cell) if cell is not None else "" for cell in row])


def upload_file(drive_service, file_path: Path, folder_id: str, mime_type: str) -> str:
    """Upload a file to Drive, replacing any existing file with the same name."""
    from googleapiclient.http import MediaFileUpload
    media = MediaFileUpload(str(file_path), mimetype=mime_type, resumable=True)
    existing = drive_service.files().list(
        q=f"name='{file_path.name}' and '{folder_id}' in parents and trashed=false",
        fields="files(id)"
    ).execute().get("files", [])
    if existing:
        uploaded = drive_service.files().update(
            fileId=existing[0]["id"], media_body=media, fields="id, name, webViewLink"
        ).execute()
    else:
        uploaded = drive_service.files().create(
            body={"name": file_path.name, "parents": [folder_id]},
            media_body=media, fields="id, name, webViewLink"
        ).execute()
    return uploaded.get("webViewLink", "")


def run(demographics_csv: Path, output_dir: Path, date_str: str = None) -> str:
    """
    Transform demographics CSV, save as Excel, derive CSV from Excel,
    upload XLSX to Drive and CSV to the 'CSV' subfolder.
    """
    if date_str is None:
        date_str = datetime.now().strftime("%Y%m%d")

    print(f"  [demographics] Transforming data...")
    rows = transform(demographics_csv)
    print(f"  [demographics] {len(rows)} unique employees found")

    # 1. Save as XLSX — all columns Text so leading zeros are preserved
    xlsx_filename = f"{date_str}_pretax_demographics.xlsx"
    xlsx_path     = output_dir / xlsx_filename
    save_as_xlsx(rows, xlsx_path)
    print(f"  [demographics] Saved: {xlsx_filename}")

    # 2. Derive CSV from the Excel file (values read back exactly as stored)
    csv_filename  = f"{date_str}_pretax_demographics.csv"
    csv_path      = output_dir / csv_filename
    xlsx_to_csv(xlsx_path, csv_path)
    print(f"  [demographics] Saved CSV from Excel: {csv_filename}")

    # 3. Upload to Drive
    print(f"  [demographics] Uploading to Google Drive...")
    drive_svc    = get_drive_service()
    XLSX_MIME    = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    subfolder_id = get_or_create_subfolder(drive_svc, GDRIVE_FOLDER_ID, "Demographics Upload File")
    xlsx_link    = upload_file(drive_svc, xlsx_path, subfolder_id, XLSX_MIME)
    print(f"  [demographics] ✓ Uploaded: {xlsx_filename}  →  {xlsx_link}")

    csv_sub_id   = get_or_create_subfolder(drive_svc, subfolder_id, "CSV")
    csv_link     = upload_file(drive_svc, csv_path, csv_sub_id, "text/csv")
    print(f"  [demographics] ✓ Uploaded CSV: {csv_filename}  →  {csv_link}")

    return xlsx_link


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        today = datetime.now().strftime("%Y-%m-%d")
        csv_path = Path.home() / "employee_navigator_reports" / f"demographic_{today}.csv"
    else:
        csv_path = Path(sys.argv[1])

    if not csv_path.exists():
        print(f"ERROR: File not found: {csv_path}")
        sys.exit(1)

    output_dir = Path.home() / "employee_navigator_reports"
    run(csv_path, output_dir, datetime.now().strftime("%Y%m%d"))
