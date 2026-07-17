#!/usr/bin/env python3
"""
Transforms raw Employee Navigator HSA and FSA CSV reports into the
pretax elections upload format and uploads to Google Drive.

Output filename: YYYYMMDD_pretax_elections.csv

Column mapping:
  Employee ID, Employee Email, Plan Year, Account Type,
  Employee Pay Period Election, Employer Pay Period Election, Account Status,
  Account Start Date, Account End Date, Employee Election, Employer Election,
  Coverage Tier
"""

import csv
import os
from datetime import datetime
from pathlib import Path
from typing import List, Optional

GDRIVE_FOLDER_ID = os.environ.get("GDRIVE_FOLDER_ID", "1L4qIfD4bha6oZpZTe7s5LbNcgRAgqLlL")
GOOGLE_CREDS     = os.environ.get("GOOGLE_CREDS_FILE", str(Path.home() / ".config/en_report/google_creds.json"))
TOKEN_FILE       = str(Path(GOOGLE_CREDS).parent / "google_token.json")
SCOPES           = [
    "https://www.googleapis.com/auth/drive.file",
]

# Columns with leading zeros that must be preserved as text in Excel
TEXT_COLUMNS = {"Employee ID"}

PAYS_PER_YEAR    = 24          # semi-monthly pay schedule
PLAN_YEAR_START  = datetime(2026, 1, 1).date()

# Maximum allowed Employee Election per coverage tier (HSA only)
# (Total annual limit minus fixed Employer Election)
HSA_EE_ELECTION_CAP = {
    "Family":     6_350.00,   # $8,750 total − $2,400 employer
    "Individual": 3_200.00,   # $4,400 total − $1,200 employer
}

# Maximum annual Employee Election by account type (FSA / DCFSA)
FSA_EE_ELECTION_CAP = {
    "FSA":   3_116.00,
    "DCFSA": 6_875.00,
}

OUTPUT_HEADERS = [
    "Employee ID",
    "Employee Email",
    "Plan Year",
    "Account Type",
    "Employee Pay Period Election",
    "Employer Pay Period Election",
    "Account Status",
    "Account Start Date",
    "Account End Date",
    "Employee Election",
    "Employer Election",
    "Coverage Tier",
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def parse_dollar(s: str) -> float:
    """Parse EN dollar strings like '$1,200.00' or '$0.00' into a float."""
    return float(s.strip().replace("$", "").replace(",", "") or 0)


def fmt_dollar(amount: float) -> str:
    """Format a float as '$1,234.56'."""
    return f"${amount:,.2f}"


def parse_date(s: str) -> Optional[datetime.date]:
    """Parse MM/DD/YYYY date strings from EN; return None if blank."""
    s = s.strip()
    if not s:
        return None
    return datetime.strptime(s, "%m/%d/%Y").date()


def termination_cutoff(today: "datetime.date") -> "datetime.date":
    """
    Returns the 21st of the previous month.
    Employees with End Date on or before this date are excluded.
    e.g. today = May 6 → cutoff = April 21
    """
    import datetime as _dt
    if today.month == 1:
        return _dt.date(today.year - 1, 12, 21)
    return _dt.date(today.year, today.month - 1, 21)


def fix_employee_id(emp_id: str, first: str, last: str) -> str:
    """Ensure Brett Kempker's Employee ID always starts with '000'."""
    if first.strip().lower() == "brett" and last.strip().lower() == "kempker":
        if not emp_id.startswith("000"):
            emp_id = "000" + emp_id
    return emp_id


def account_start_date(en_start: Optional[datetime.date]) -> str:
    """Return max(en_start, PLAN_YEAR_START) in YYYY-MM-DD format."""
    if en_start is None or en_start <= PLAN_YEAR_START:
        return PLAN_YEAR_START.strftime("%Y-%m-%d")
    return en_start.strftime("%Y-%m-%d")


# ── HSA transform ─────────────────────────────────────────────────────────────

def transform_hsa(hsa_csv: Path, today: datetime.date) -> List[dict]:
    """
    Read the HSA report and return election rows.

    Coverage Tier logic:
      Yearly Employer Contribution >= $1,500  →  Family  (standard $2,400 or prorated ~$2,000)
      Yearly Employer Contribution  < $1,500  →  Individual (standard $1,200 or prorated ~$1,000)

    Employer Pay Period Election:
      Family → $100.00  (fixed)
      Individual → $50.00 (fixed)

    Employer Pay Period for already-terminated employees (End Date <= today) → $0.00
    """
    cutoff = termination_cutoff(today)
    rows = []

    with open(hsa_csv, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            emp_id    = row.get("Employee ID", "").strip()
            first     = row.get("First Name", "").strip()
            last      = row.get("Last Name", "").strip()
            email     = row.get("Work Email", "").strip()
            start_raw = row.get("Start Date", "").strip()
            end_raw   = row.get("End Date", "").strip()

            if not emp_id or not email:
                continue

            emp_id = fix_employee_id(emp_id, first, last)

            end_date   = parse_date(end_raw)
            start_date = parse_date(start_raw)

            # Skip employees terminated more than CUTOFF_DAYS ago
            if end_date and end_date < cutoff:
                continue

            yearly_employer = parse_dollar(row.get("Yearly Employer Contribution", "0"))
            ee_per_pay      = parse_dollar(row.get("Employee Per Pay", "0"))

            # Coverage Tier from annual employer contribution
            coverage_tier = "Family" if yearly_employer >= 1500 else "Individual"

            # Employer Pay Period: $0 if already terminated, else fixed Family/Individual rate
            if end_date and end_date <= today:
                er_per_pay_out = 0.0
            else:
                er_per_pay_out = 100.0 if coverage_tier == "Family" else 50.0

            status = "Terminated" if end_date else "Active"

            rows.append({
                "Employee ID":                 emp_id,
                "Employee Email":              email,
                "Plan Year":                   "",
                "Account Type":                "HSA",
                "Employee Pay Period Election": fmt_dollar(ee_per_pay),
                "Employer Pay Period Election": fmt_dollar(er_per_pay_out),
                "Account Status":              status,
                "Account Start Date":          account_start_date(start_date),
                "Account End Date":            "",
                "Employee Election":           fmt_dollar(min(ee_per_pay * PAYS_PER_YEAR, HSA_EE_ELECTION_CAP[coverage_tier])),
                "Employer Election":           fmt_dollar(yearly_employer),
                "Coverage Tier":               coverage_tier,
            })

    return rows


# ── FSA / DCFSA transform ─────────────────────────────────────────────────────

def transform_fsa(fsa_csv: Path, today: datetime.date) -> List[dict]:
    """
    Read the FSA report and return FSA + DCFSA election rows.

    Rules:
    - Skip rows where EE Per Pay Cost = $0 (declined / not enrolled)
    - Skip rows where Employee ID is blank
    - For each (Employee ID, Account Type), keep only the most recent
      active enrollment: End Date blank OR End Date >= (today - CUTOFF_DAYS)
    - Account Type: "FSA" for Medical FSA, "DCFSA" for Dependent Care
    - Coverage Tier: FSA → "Individual", DCFSA → "Family"
    - Employer amounts always $0.00 for FSA/DCFSA
    - Account Status: "Active" if End Date blank, "Terminated" if End Date present
    - Employer Pay Period → $0 if already terminated (End Date <= today)
    """
    cutoff = termination_cutoff(today)

    # Collect candidate rows: (employee_id, account_type) → list of rows
    from collections import defaultdict
    candidates = defaultdict(list)

    with open(fsa_csv, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            emp_id = row.get("Employee ID", "").strip()
            first  = row.get("First Name", "").strip()
            last   = row.get("Last Name", "").strip()
            email  = row.get("Work Email", "").strip()

            if not emp_id or not email:
                continue

            emp_id = fix_employee_id(emp_id, first, last)

            ee_per_pay = parse_dollar(row.get("EE Per Pay Cost", "0"))
            if ee_per_pay == 0:
                continue  # not enrolled / declined

            plan = row.get("Plan", "")
            if "Dependent Care" in plan:
                account_type  = "DCFSA"
                coverage_tier = "Family"
            else:
                account_type  = "FSA"
                coverage_tier = "Individual"

            start_raw = row.get("Start Date", "").strip()
            end_raw   = row.get("End Date", "").strip()
            end_date  = parse_date(end_raw)
            start_date = parse_date(start_raw)

            # Skip if ended too long ago
            if end_date and end_date < cutoff:
                continue

            candidates[(emp_id, account_type)].append({
                "emp_id":        emp_id,
                "email":         email,
                "account_type":  account_type,
                "coverage_tier": coverage_tier,
                "ee_per_pay":    ee_per_pay,
                "start_date":    start_date,
                "end_date":      end_date,
            })

    rows = []
    for (emp_id, account_type), cands in candidates.items():
        # Pick the enrollment with the latest Start Date
        best = max(cands, key=lambda c: c["start_date"] or PLAN_YEAR_START)

        end_date   = best["end_date"]
        start_date = best["start_date"]
        ee_per_pay = best["ee_per_pay"]
        coverage_tier = best["coverage_tier"]
        email      = best["email"]

        status = "Terminated" if end_date else "Active"

        # Employer is always $0 for FSA/DCFSA
        rows.append({
            "Employee ID":                 emp_id,
            "Employee Email":              email,
            "Plan Year":                   str(PLAN_YEAR_START.year),
            "Account Type":                account_type,
            "Employee Pay Period Election": fmt_dollar(ee_per_pay),
            "Employer Pay Period Election": "$0.00",
            "Account Status":              status,
            "Account Start Date":          account_start_date(start_date),
            "Account End Date":            "",
            "Employee Election":           fmt_dollar(min(ee_per_pay * PAYS_PER_YEAR, FSA_EE_ELECTION_CAP[account_type])),
            "Employer Election":           "$0.00",
            "Coverage Tier":               coverage_tier,
        })

    return rows


# ── Sort and deduplicate ───────────────────────────────────────────────────────

def sort_rows(rows: List[dict]) -> List[dict]:
    """Sort: HSA first, then DCFSA, then FSA; within each by Employee ID."""
    type_order = {"HSA": 0, "DCFSA": 1, "FSA": 2}
    return sorted(rows, key=lambda r: (
        type_order.get(r["Account Type"], 9),
        r["Employee ID"],
    ))


# ── Google Drive ──────────────────────────────────────────────────────────────

def get_drive_service():
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build

    creds = None
    if os.path.exists(TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(GOOGLE_CREDS, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, "w") as f:
            f.write(creds.to_json())
    return build("drive", "v3", credentials=creds)


def get_or_create_subfolder(service, parent_id: str, folder_name: str) -> str:
    query = (
        f"name='{folder_name}' and "
        f"'{parent_id}' in parents and "
        "mimeType='application/vnd.google-apps.folder' and "
        "trashed=false"
    )
    results = service.files().list(q=query, fields="files(id, name)").execute()
    files = results.get("files", [])
    if files:
        return files[0]["id"]
    metadata = {
        "name": folder_name,
        "mimeType": "application/vnd.google-apps.folder",
        "parents": [parent_id],
    }
    folder = service.files().create(body=metadata, fields="id").execute()
    return folder["id"]


def save_as_xlsx(rows: List[dict], xlsx_path: Path):
    """Save rows as Excel (.xlsx) with ALL columns set to Text format to prevent auto-conversion."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for col_idx, header in enumerate(OUTPUT_HEADERS, 1):
        ws.cell(row=1, column=col_idx, value=header)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, header in enumerate(OUTPUT_HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(row.get(header, "")))
            cell.number_format = "@"  # Text format on ALL columns
    wb.save(str(xlsx_path))


def xlsx_to_csv(xlsx_path: Path, csv_path: Path):
    """
    Read the saved XLSX (all cells stored as text) and export as UTF-8 BOM CSV.
    Values are taken directly from Excel's stored cell values so leading zeros,
    employee IDs, etc. are preserved exactly as they appear in the Excel file.
    """
    from openpyxl import load_workbook
    wb = load_workbook(str(xlsx_path), data_only=True)
    ws = wb.active
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([str(cell) if cell is not None else "" for cell in row])


def upload_file(drive_service, file_path: Path, folder_id: str, mime_type: str) -> str:
    """Upload a file to Drive, replacing any existing file with the same name."""
    from googleapiclient.http import MediaFileUpload
    media = MediaFileUpload(str(file_path), mimetype=mime_type, resumable=True)
    existing = drive_service.files().list(
        q=f"name='{file_path.name}' and '{folder_id}' in parents and trashed=false",
        fields="files(id)"
    ).execute().get("files", [])
    if existing:
        uploaded = drive_service.files().update(
            fileId=existing[0]["id"], media_body=media, fields="id, name, webViewLink"
        ).execute()
    else:
        uploaded = drive_service.files().create(
            body={"name": file_path.name, "parents": [folder_id]},
            media_body=media, fields="id, name, webViewLink"
        ).execute()
    return uploaded.get("webViewLink", "")


# ── Public entry point ────────────────────────────────────────────────────────

def run(hsa_csv: Path, fsa_csv: Path, output_dir: Path, date_str: str = None) -> str:
    """
    Transform HSA + FSA CSVs into a combined elections upload CSV and upload to Drive.
    Returns the Google Drive link.
    """
    if date_str is None:
        date_str = datetime.now().strftime("%Y%m%d")

    today = datetime.now().date()

    print(f"  [elections] Processing HSA data...")
    hsa_rows = transform_hsa(hsa_csv, today)
    print(f"  [elections] {len(hsa_rows)} HSA rows")

    print(f"  [elections] Processing FSA/DCFSA data...")
    fsa_rows = transform_fsa(fsa_csv, today)
    print(f"  [elections] {len(fsa_rows)} FSA/DCFSA rows")

    all_rows = sort_rows(hsa_rows + fsa_rows)

    # 1. Save as XLSX — all columns set to Text so nothing is auto-converted
    xlsx_filename = f"{date_str}_pretax_elections.xlsx"
    xlsx_path     = output_dir / xlsx_filename
    save_as_xlsx(all_rows, xlsx_path)
    print(f"  [elections] Saved: {xlsx_filename}")

    # 2. Derive CSV from the Excel file (values read back exactly as stored)
    csv_filename  = f"{date_str}_pretax_elections.csv"
    csv_path      = output_dir / csv_filename
    xlsx_to_csv(xlsx_path, csv_path)
    print(f"  [elections] Saved CSV from Excel: {csv_filename} ({len(all_rows)} rows total)")

    # 3. Upload to Drive
    print(f"  [elections] Uploading to Google Drive...")
    service      = get_drive_service()
    XLSX_MIME    = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    subfolder_id = get_or_create_subfolder(service, GDRIVE_FOLDER_ID, "Elections Upload File")
    xlsx_link    = upload_file(service, xlsx_path, subfolder_id, XLSX_MIME)
    print(f"  [elections] ✓ Uploaded: {xlsx_filename}  →  {xlsx_link}")

    csv_sub_id   = get_or_create_subfolder(service, subfolder_id, "CSV")
    csv_link     = upload_file(service, csv_path, csv_sub_id, "text/csv")
    print(f"  [elections] ✓ Uploaded CSV: {csv_filename}  →  {csv_link}")

    return xlsx_link


if __name__ == "__main__":
    import sys
    today_str = datetime.now().strftime("%Y-%m-%d")
    date_str  = datetime.now().strftime("%Y%m%d")

    if len(sys.argv) >= 3:
        hsa_path = Path(sys.argv[1])
        fsa_path = Path(sys.argv[2])
    else:
        reports_dir = Path.home() / "employee_navigator_reports"
        hsa_path = reports_dir / f"hsa_{today_str}.csv"
        fsa_path = reports_dir / f"fsa_{today_str}.csv"

    for p in (hsa_path, fsa_path):
        if not p.exists():
            print(f"ERROR: File not found: {p}")
            sys.exit(1)

    output_dir = Path.home() / "employee_navigator_reports"
    run(hsa_path, fsa_path, output_dir, date_str)
